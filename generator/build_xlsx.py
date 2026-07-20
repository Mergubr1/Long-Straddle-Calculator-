"""
Generate the Long Straddle Calculator as a fully formula-driven .xlsx workbook.

Usage:
    python build_xlsx.py [-o OUTPUT_PATH] [--symbol XYZ]

All strategy math (breakevens, P/L, payoff table, chart) is written as live
Excel formulas -- editing any blue input cell recalculates the whole sheet.
"""
import argparse
import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation


def build(output_path: str = "Long_Straddle_Calculator.xlsx", symbol: str = "XYZ") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Long Straddle Calculator"

    FONT_NAME = "Arial"

    # ---------- Styles ----------
    TITLE_FONT   = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
    SUB_FONT     = Font(name=FONT_NAME, size=10, italic=True, color="FFFFFF")
    SECTION_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    LABEL_FONT   = Font(name=FONT_NAME, size=10)
    NOTE_FONT    = Font(name=FONT_NAME, size=9, italic=True, color="666666")
    INPUT_FONT   = Font(name=FONT_NAME, size=10, bold=True, color="0000FF")
    FORMULA_FONT = Font(name=FONT_NAME, size=10, bold=True, color="000000")
    HEADER_TABLE_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")

    TITLE_FILL   = PatternFill("solid", fgColor="1F4E78")
    SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
    INPUT_FILL   = PatternFill("solid", fgColor="FFF2CC")
    OUTPUT_FILL  = PatternFill("solid", fgColor="E2EFDA")
    TABLE_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")

    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    CUR2 = '$#,##0.00;($#,##0.00);"-"'
    CUR0 = '$#,##0;($#,##0);"-"'
    PCT1 = '0.0%;(0.0%);"-"'
    DATE_FMT = 'mm/dd/yyyy'

    def set_cell(row, col, value, font=None, fill=None, fmt=None, align=None, border=True, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        if font: c.font = font
        if fill: c.fill = fill
        if fmt: c.number_format = fmt
        c.alignment = align or Alignment(vertical="center", wrap_text=wrap)
        if border: c.border = BORDER
        return c

    # Column widths
    widths = {'A':2,'B':22,'C':14,'D':32,'E':2,'F':30,'G':16,'H':40}
    for col,w in widths.items():
        ws.column_dimensions[col].width = w

    # ---------- Title ----------
    ws.merge_cells('A1:H1')
    set_cell(1,1,"LONG STRADDLE CALCULATOR — BUY CALL + BUY PUT", TITLE_FONT, TITLE_FILL, align=Alignment(horizontal="left", vertical="center", indent=1), border=False)
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:H2')
    set_cell(2,1,"Buy one call and one put with the same strike price and expiration date. Blue cells are editable inputs.",
             SUB_FONT, TITLE_FILL, align=Alignment(horizontal="left", vertical="center", indent=1), border=False)
    ws.row_dimensions[2].height = 18

    # ---------- Section headers ----------
    ws.merge_cells('B4:D4')
    set_cell(4,2,"EDITABLE TRADE INPUTS", SECTION_FONT, SECTION_FILL, align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.merge_cells('F4:H4')
    set_cell(4,6,"AUTOMATIC STRATEGY RESULTS", SECTION_FONT, SECTION_FILL, align=Alignment(horizontal="left", vertical="center", indent=1))

    # ---------- Inputs (col B label, C value, D note) ----------
    inputs = [
        ("Underlying Symbol", symbol, "User-editable symbol", "text"),
        ("Stock Price at Entry ($)", 50.00, "Current share price when entering the trade", CUR2),
        ("Strike Price ($)", 50.00, "Same strike for the call and put", CUR2),
        ("Call Premium Paid ($/share)", 3.00, "Price paid for the call", CUR2),
        ("Put Premium Paid ($/share)", 2.00, "Price paid for the put", CUR2),
        ("Number of Straddles", 1, "One straddle = one call contract + one put contract", '0'),
        ("Shares per Contract", 100, "Standard U.S. equity option multiplier", '0'),
        ("Expiration Date", None, "Same expiration for both option legs", DATE_FMT),
        ("Days to Expiration", None, "Automatically calculated from the expiration date", '0'),
        ("Stock Price at Expiration ($)", 60.00, "Change this input to test a specific outcome", CUR2),
        ("Fees & Commissions ($)", 0, "Total estimated opening and closing costs", CUR2),
        ("Implied Volatility Outlook", "Expected to Rise", "Rising volatility generally helps a long straddle", "text"),
        ("Scenario Low (% of Strike)", 0.50, "Lowest price shown in the payoff table", PCT1),
        ("Scenario High (% of Strike)", 1.50, "Highest price shown in the payoff table", PCT1),
        ("Scenario Intervals", 20, "Creates intervals + 1 payoff points", '0'),
    ]

    row0 = 5
    rows = {}
    for i,(label,val,note,fmt) in enumerate(inputs):
        r = row0+i
        rows[label]=r
        set_cell(r,2,label, LABEL_FONT, None)
        if label=="Expiration Date":
            val = datetime.date(2026,9,18)
            c = set_cell(r,3,val, INPUT_FONT, INPUT_FILL, fmt)
        elif label=="Days to Expiration":
            c = set_cell(r,3,f"=C{rows['Expiration Date']}-TODAY()", FORMULA_FONT, OUTPUT_FILL, fmt)
        else:
            c = set_cell(r,3,val, INPUT_FONT, INPUT_FILL, fmt)
        set_cell(r,4,note, NOTE_FONT, None, align=Alignment(vertical="center", wrap_text=True), border=False)

    # Data validation for IV outlook dropdown
    dv = DataValidation(type="list", formula1='"Expected to Rise,Expected to Fall,Expected to Stay Flat"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=rows["Implied Volatility Outlook"], column=3))

    R = rows  # shortcut
    SYM   = f"C{R['Underlying Symbol']}"
    S0    = f"C{R['Stock Price at Entry ($)']}"
    K     = f"C{R['Strike Price ($)']}"
    CALLP = f"C{R['Call Premium Paid ($/share)']}"
    PUTP  = f"C{R['Put Premium Paid ($/share)']}"
    NUM   = f"C{R['Number of Straddles']}"
    MULT  = f"C{R['Shares per Contract']}"
    EXPD  = f"C{R['Expiration Date']}"
    DTE   = f"C{R['Days to Expiration']}"
    SEXP  = f"C{R['Stock Price at Expiration ($)']}"
    FEES  = f"C{R['Fees & Commissions ($)']}"
    IVOUT = f"C{R['Implied Volatility Outlook']}"
    SCLOW = f"C{R['Scenario Low (% of Strike)']}"
    SCHIGH= f"C{R['Scenario High (% of Strike)']}"
    SCINT = f"C{R['Scenario Intervals']}"

    # ---------- Outputs (col F label, G value, H note) ----------
    outputs = [
        ("Combined Premium ($/share)", f"={CALLP}+{PUTP}", "Call premium + put premium", CUR2),
        ("Total Debit / Cost ($)", f"=(({CALLP}+{PUTP})*{MULT}*{NUM})+{FEES}", "Maximum loss, including fees", CUR0),
        ("Lower Breakeven ($)", f"={K}-({CALLP}+{PUTP})", "Strike \u2212 combined premium", CUR2),
        ("Upper Breakeven ($)", f"={K}+({CALLP}+{PUTP})", "Strike + combined premium", CUR2),
        ("Call Intrinsic Value ($/share)", f"=MAX({SEXP}-{K},0)", "MAX(expiration price \u2212 strike, 0)", CUR2),
        ("Put Intrinsic Value ($/share)", f"=MAX({K}-{SEXP},0)", "MAX(strike \u2212 expiration price, 0)", CUR2),
        ("Combined Value at Expiration ($/share)", None, "Call value + put value", CUR2),
        ("Gross Profit / Loss ($)", None, "Before fees", CUR0),
        ("Net Profit / Loss ($)", None, "After fees", CUR0),
        ("Return on Debit (%)", None, "Net P/L \u00f7 total debit", PCT1),
        ("Expiration Result", None, "Profit, breakeven, or loss", "text"),
        ("Maximum Profit", None, "Unlimited upside; substantial downside potential", "text"),
        ("Maximum Loss ($)", None, "Limited to total debit", CUR0),
        ("Best Market Outlook", None, "A large move in either direction", "text"),
        ("Time Decay Effect", None, "Hurts the position", "text"),
    ]

    orow0 = 5
    orows = {}
    for i,(label,formula,note,fmt) in enumerate(outputs):
        r = orow0+i
        orows[label]=r
        set_cell(r,6,label, LABEL_FONT, None, align=Alignment(vertical="center", wrap_text=True))
        fnt = FORMULA_FONT if fmt!="text" else Font(name=FONT_NAME, size=10, bold=True)
        c = set_cell(r,7,formula if formula else "", fnt, OUTPUT_FILL, fmt if fmt!="text" else None,
                     align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        set_cell(r,8,note, NOTE_FONT, None, align=Alignment(vertical="center", wrap_text=True), border=False)

    O = orows
    COMBPREM = f"G{O['Combined Premium ($/share)']}"
    TOTDEBIT = f"G{O['Total Debit / Cost ($)']}"
    LOWBE    = f"G{O['Lower Breakeven ($)']}"
    UPBE     = f"G{O['Upper Breakeven ($)']}"
    CALLINT  = f"G{O['Call Intrinsic Value ($/share)']}"
    PUTINT   = f"G{O['Put Intrinsic Value ($/share)']}"
    COMBVAL  = f"G{O['Combined Value at Expiration ($/share)']}"
    GROSSPL  = f"G{O['Gross Profit / Loss ($)']}"
    NETPL    = f"G{O['Net Profit / Loss ($)']}"
    RETDEB   = f"G{O['Return on Debit (%)']}"
    EXPRES   = f"G{O['Expiration Result']}"
    MAXPROF  = f"G{O['Maximum Profit']}"
    MAXLOSS  = f"G{O['Maximum Loss ($)']}"
    BESTOUT  = f"G{O['Best Market Outlook']}"
    TDECAY   = f"G{O['Time Decay Effect']}"

    ws[COMBVAL] = f"={CALLINT}+{PUTINT}"
    ws[GROSSPL] = f"=({COMBVAL}-{COMBPREM})*{MULT}*{NUM}"
    ws[NETPL]   = f"={GROSSPL}-{FEES}"
    ws[RETDEB]  = f"=IF({TOTDEBIT}=0,\"\",{NETPL}/{TOTDEBIT})"
    ws[EXPRES]  = f'=IF({NETPL}>0,"PROFIT",IF({NETPL}<0,"LOSS","BREAKEVEN"))'
    ws[MAXPROF] = f'="Unlimited above "&TEXT({UPBE},"$#,##0.00")&"; downside gains increase below "&TEXT({LOWBE},"$#,##0.00")'
    ws[MAXLOSS] = f"={TOTDEBIT}"
    ws[BESTOUT] = f'="Large move above "&TEXT({UPBE},"$#,##0.00")&" or below "&TEXT({LOWBE},"$#,##0.00")'
    ws[TDECAY]  = '="Negative \u2014 both purchased options lose time value"'

    # Conditional formatting for Expiration Result cell
    ws.conditional_formatting.add(EXPRES,
        FormulaRule(formula=[f'{EXPRES}="PROFIT"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=FONT_NAME, bold=True, color="006100")))
    ws.conditional_formatting.add(EXPRES,
        FormulaRule(formula=[f'{EXPRES}="LOSS"'], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(name=FONT_NAME, bold=True, color="9C0006")))
    ws.conditional_formatting.add(EXPRES,
        FormulaRule(formula=[f'{EXPRES}="BREAKEVEN"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=FONT_NAME, bold=True, color="9C6500")))

    # ---------- Payoff table ----------
    table_title_row = max(row0+len(inputs), orow0+len(outputs)) + 2
    ws.merge_cells(f'A{table_title_row}:H{table_title_row}')
    set_cell(table_title_row,1,"EXPIRATION PAYOFF TABLE", SECTION_FONT, SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[table_title_row].height = 20

    note_row = table_title_row+1
    ws.merge_cells(f'A{note_row}:H{note_row}')
    set_cell(note_row,1,"Prices scale automatically with Strike Price, Scenario Low/High %, and Scenario Intervals above. Rows beyond the chosen interval count are left blank.",
             NOTE_FONT, None, align=Alignment(horizontal="left", vertical="center", indent=1), border=False)

    header_row = note_row+1
    headers = ["Stock at Expiration ($)","Call Intrinsic ($/share)","Put Intrinsic ($/share)",
               "Combined Value ($/share)","Net P/L ($/share)","Total Net P/L ($)","Return on Debit (%)"]
    col_map = {0:2,1:3,2:4,3:5,4:6,5:7,6:8}  # B..H
    for i,h in enumerate(headers):
        set_cell(header_row, col_map[i], h, HEADER_TABLE_FONT, TABLE_HEADER_FILL,
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[header_row].height = 30

    MAX_ROWS = 40  # supports scenario intervals up to 40
    data_start = header_row+1
    for i in range(MAX_ROWS+1):
        r = data_start+i
        n = i  # row index 0..MAX_ROWS
        price_formula = (f'=IF({n}>{SCINT},"",{K}*{SCLOW}+{n}*({K}*{SCHIGH}-{K}*{SCLOW})/{SCINT})')
        set_cell(r,2, price_formula, Font(name=FONT_NAME,size=10), None, CUR2, align=Alignment(horizontal="right"))
        pcell = f"B{r}"
        call_f = f'=IF({pcell}="","",MAX({pcell}-{K},0))'
        put_f  = f'=IF({pcell}="","",MAX({K}-{pcell},0))'
        set_cell(r,3, call_f, Font(name=FONT_NAME,size=10), None, CUR2, align=Alignment(horizontal="right"))
        set_cell(r,4, put_f, Font(name=FONT_NAME,size=10), None, CUR2, align=Alignment(horizontal="right"))
        comb_f = f'=IF({pcell}="","",C{r}+D{r})'
        set_cell(r,5, comb_f, Font(name=FONT_NAME,size=10), None, CUR2, align=Alignment(horizontal="right"))
        netps_f = f'=IF({pcell}="","",E{r}-{COMBPREM})'
        set_cell(r,6, netps_f, Font(name=FONT_NAME,size=10), None, CUR2, align=Alignment(horizontal="right"))
        total_f = f'=IF({pcell}="","",(F{r}*{MULT}*{NUM})-{FEES})'
        set_cell(r,7, total_f, Font(name=FONT_NAME,size=10,bold=True), None, CUR0, align=Alignment(horizontal="right"))
        ret_f = f'=IF({pcell}="","",IF({TOTDEBIT}=0,"",G{r}/{TOTDEBIT}))'
        set_cell(r,8, ret_f, Font(name=FONT_NAME,size=10), None, PCT1, align=Alignment(horizontal="right"))
        if i % 2 == 1:
            for cc in range(2,9):
                ws.cell(row=r, column=cc).fill = ALT_ROW_FILL

    data_end = data_start+MAX_ROWS

    # Conditional formatting for Total Net P/L column (green positive / red negative)
    totpl_range = f"G{data_start}:G{data_end}"
    ws.conditional_formatting.add(totpl_range,
        CellIsRule(operator='greaterThan', formula=['0'], font=Font(name=FONT_NAME, bold=True, color="006100")))
    ws.conditional_formatting.add(totpl_range,
        CellIsRule(operator='lessThan', formula=['0'], font=Font(name=FONT_NAME, bold=True, color="9C0006")))

    # Freeze panes below the header block
    ws.freeze_panes = f"B{header_row+1}"

    # ---------- Chart ----------
    chart = LineChart()
    chart.title = "Long Straddle Payoff at Expiration"
    chart.style = 2
    chart.y_axis.title = 'Total Net P/L ($)'
    chart.x_axis.title = 'Stock Price at Expiration ($)'
    chart.height = 9
    chart.width = 22

    data = Reference(ws, min_col=7, min_row=header_row, max_row=data_end)
    cats = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[0].graphicalProperties.line.solidFill = "1F4E78"
    chart.series[0].smooth = False

    ws.add_chart(chart, f"J{table_title_row}")

    # ---------- Assumptions / documentation note ----------
    doc_row = data_end + 2
    ws.merge_cells(f'A{doc_row}:H{doc_row}')
    set_cell(doc_row,1,"ASSUMPTIONS & NOTES", SECTION_FONT, SECTION_FILL,
             align=Alignment(horizontal="left", vertical="center", indent=1))
    notes = [
     "• All example figures (symbol, prices, premiums, dates) are illustrative placeholders entered by the user — replace the blue cells with your own trade data.",
     "• This model values each leg at expiration using intrinsic value only (MAX formulas); it does not price the options before expiration (no Black-Scholes / time-value estimate).",
     "• \"Number of Straddles\" scales both the call and put legs equally, consistent with a standard long straddle (1 call + 1 put per straddle).",
     "• Fees & Commissions are treated as a single total figure applied once to the overall position, not per contract.",
     "• The payoff table supports up to 40 scenario intervals; increase 'Scenario Intervals' beyond 40 and additional rows will need to be added manually.",
    ]
    for i,n in enumerate(notes):
        r = doc_row+1+i
        ws.merge_cells(f'A{r}:H{r}')
        set_cell(r,1,n, NOTE_FONT, None, align=Alignment(horizontal="left", vertical="center", wrap_text=True), border=False)

    ws.sheet_view.showGridLines = False

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate the Long Straddle Calculator workbook.")
    parser.add_argument("-o", "--output", default="Long_Straddle_Calculator.xlsx",
                         help="Output .xlsx path (default: Long_Straddle_Calculator.xlsx)")
    parser.add_argument("--symbol", default="XYZ", help="Underlying ticker symbol (default: XYZ)")
    args = parser.parse_args()

    path = build(output_path=args.output, symbol=args.symbol)
    print(f"Saved: {path}")
    print("Tip: open in Excel/LibreOffice -- all cells are live formulas.")


if __name__ == "__main__":
    main()
